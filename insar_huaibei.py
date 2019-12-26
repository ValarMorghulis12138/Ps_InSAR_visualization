import os
from qgis.core import *
try:
    from osgeo import gdal
    from osgeo import ogr
    from osgeo import osr
except ImportError:
    import gdal
    import ogr
    import osr
import openpyxl
import xml.dom.minidom as xdm
import imageio
import report_generator


def replace_shpsource(shp_path, out_xml):
    # shp_path.replace('\\', '/')
    dom = xdm.parse('huaibei.qgs')
    root = dom.documentElement
    root.getElementsByTagName('datasource')[0].firstChild.data = shp_path
    with open(out_xml, 'w', encoding='utf-8') as fh:
        dom.writexml(fh, encoding='utf-8')


def replace_ratsource(rat_path, out_xml):
    # shp_path.replace('\\', '/')
    dom = xdm.parse('huaibei_raster.qgs')
    root = dom.documentElement
    root.getElementsByTagName('datasource')[0].firstChild.data = rat_path
    with open(out_xml, 'w', encoding='utf-8') as fh:
        dom.writexml(fh, encoding='utf-8')


def xlsx2shp(in_xlsx, out_shp):
    wb = openpyxl.load_workbook(in_xlsx)
    ws = wb.get_sheet_by_name('Sheet1')
    gdal.AllRegister()
    # 为了支持中文路径，请添加下面这句代码
    gdal.SetConfigOption("GDAL_FILENAME_IS_UTF8", "NO")
    # 为了使属性表字段支持中文，请添加下面这句
    gdal.SetConfigOption("SHAPE_ENCODING", "CP936")
    # 注册所有的驱动
    ogr.RegisterAll()
    # set up the shapefile driver
    driver = ogr.GetDriverByName("ESRI Shapefile")
    # create the data source
    data_source = driver.CreateDataSource(out_shp)
    # create the spatial reference, WGS84
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)
    # create the layer
    layer = data_source.CreateLayer("psdata", srs, ogr.wkbPoint)
    # Add the fields we're interested in
    layer.CreateField(ogr.FieldDefn("Latitude", ogr.OFTReal))
    layer.CreateField(ogr.FieldDefn("Longitude", ogr.OFTReal))
    layer.CreateField(ogr.FieldDefn("Value", ogr.OFTReal))
    # Process the xlsx file and add the attributes and features to the shpfile
    for i in range(1, 134935):
        lon = ws['A' + str(i)].value
        lat = ws['B' + str(i)].value
        value = ws['C' + str(i)].value
        if (lon is not None) and (lat is not None):
            # create the feature
            feature = ogr.Feature(layer.GetLayerDefn())
            # Set the attributes using the values from the delimited text file
            feature.SetField("Latitude", lat)
            feature.SetField("Longitude", lon)
            feature.SetField("Value", value)
            # create the WKT for the feature using Python string formatting
            wkt = "POINT(%f %f)" % (float(lon), float(lat))
            # Create the point from the Well Known Txt
            point = ogr.CreateGeometryFromWkt(wkt)
            # Set the feature geometry using the point
            feature.SetGeometry(point)
            # Create the feature in the layer (shapefile)
            layer.CreateFeature(feature)
            # Dereference the feature
            feature = None
    # Save and close the data source
    data_source = None


def shp2raster(in_shp, out_raster):
    # interpolate shapefile to raster using gdal.grid()
    gdal.Grid(out_raster, in_shp, format='GTiff', zfield='Value',
              algorithm='invdist:power=3.0:smoothing=0.0:min_points=12')


if __name__ == "__main__":
    input_dir = "C:/Users/codhy/Desktop/Sar/input"
    output_dir = "C:/Users/codhy/Desktop/Sar/output"

    jpg_list1 = []
    jpg_list2 = []

    # main process in qgis
    # supply path to qgis install location
    QgsApplication.setPrefixPath("C:/PROGRA~1/QGIS3~1.8/apps/qgis", True)
    # create a reference to the QgsApplication, setting the second argument to False disables the GUI
    qgs = QgsApplication([], False)
    # load providers
    qgs.initQgis()

    for f in os.listdir(input_dir):
        excel_path = os.path.join(input_dir, f)
        out_dir = os.path.join(output_dir, f.split('.')[0])
        if not os.path.exists(out_dir):
            os.mkdir(out_dir)
        # get data time from in_xlsx file name
        time_str = f.split('.')[0]
        # convert excel to shapefile
        out_shp_path = os.path.join(out_dir, time_str + '.shp')
        if not os.path.exists(out_shp_path):
            xlsx2shp(excel_path, out_shp_path)
        # convert shapefile to raster
        out_rat_path = os.path.join(out_dir, time_str + '.tif')
        if not os.path.exists(out_rat_path):
            shp2raster(out_shp_path, out_rat_path)
        # 生成沉降监测点图
        # replace shapefile datasource in qgs project file
        out_qgs_path = time_str + '_shp.qgs'
        replace_shpsource(out_shp_path, out_qgs_path)
        # export thematic map
        time_str_chinese = time_str[0:4] + u'年' + time_str[4:6] + u'月' + time_str[6:] + u'日'
        project = QgsProject.instance()
        project.read(out_qgs_path)
        manager = project.layoutManager()
        layout = manager.layoutByName('huaibei_InSar')
        # 修改layout中label内容
        label = layout.itemById('title')
        label.setText(time_str_chinese+"淮北市沉降监测图")
        exporter = QgsLayoutExporter(layout)
        out_jpg1_path = os.path.join(out_dir, time_str + '_shp.jpg')
        if os.path.exists(out_jpg1_path):
            os.remove(out_jpg1_path)
        exporter.exportToImage(out_jpg1_path, QgsLayoutExporter.ImageExportSettings())
        jpg_list1.append(out_jpg1_path)
        project = None
        os.remove(out_qgs_path)
        # 生成沉降区域图
        # replace raster datasource in qgs project file
        out_qgs_path = time_str + '_rat.qgs'
        replace_ratsource(out_rat_path, out_qgs_path)
        # export thematic map
        project = QgsProject.instance()
        project.read(out_qgs_path)
        manager = project.layoutManager()
        layout = manager.layoutByName('huaibei_InSar')
        # 修改layout中label内容
        label = layout.itemById('title')
        label.setText(time_str_chinese + "淮北市沉降明显区域图")
        exporter = QgsLayoutExporter(layout)
        out_jpg2_path = os.path.join(out_dir, time_str + '_rat.jpg')
        if os.path.exists(out_jpg2_path):
            os.remove(out_jpg2_path)
        exporter.exportToImage(out_jpg2_path, QgsLayoutExporter.ImageExportSettings())
        jpg_list2.append(out_jpg2_path)
        project = None
        os.remove(out_qgs_path)
        # 读取数据库，出统计报告
        out_doc_path = os.path.join(out_dir, 'InSAR_report_' + time_str + '.docx')
        report_generator.export(time_str=time_str, pic1_path=out_jpg1_path, pic2_path=out_jpg2_path, out_doc_path=out_doc_path)
    # 生成动图
    gif_images1 = []
    for jpg_path in jpg_list1:
        gif_images1.append(imageio.imread(jpg_path))
    gif_images2 = []
    for jpg_path in jpg_list2:
        gif_images2.append(imageio.imread(jpg_path))
    imageio.mimsave("test1.gif", gif_images1, fps=4)
    imageio.mimsave("test2.gif", gif_images2, fps=4)

    # 退出qgs
    qgs.exitQgis()
